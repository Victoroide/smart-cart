from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.db import transaction
from django.http import HttpResponse
from datetime import datetime, timedelta
from io import BytesIO
import json
import os # Importado para os.path.join si se usa directamente, aunque el modelo lo maneja
from django.conf import settings
from django.core.files.base import ContentFile
from uuid import uuid4
import logging

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from django.utils.translation import gettext

from core.models import LoggerService
from core.pagination import CustomPagination

from .models import Report
from .serializers import ReportSerializer, ReportCreateSerializer
from app.orders.models import Order, OrderItem, Payment
from app.products.models import Product, Inventory
from app.authentication.models import User # Asumiendo que User está aquí
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiExample, OpenApiResponse
from drf_spectacular.types import OpenApiTypes

logger = logging.getLogger(__name__)

@extend_schema(tags=['Report'])
class ReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CustomPagination

    @extend_schema(
        summary='Get all reports',
        description='Retrieves a paginated list of reports. Staff users can see all reports, while regular users only see their own reports.',
        responses={
            200: ReportSerializer(many=True),
            500: OpenApiResponse(description='Server error'),
        },
        parameters=[
            OpenApiParameter(
                name='page', 
                description='Page number for pagination',
                required=False, 
                type=OpenApiTypes.INT
            ),
            OpenApiParameter(
                name='page_size', 
                description='Number of items per page',
                required=False,
                type=OpenApiTypes.INT
            )
        ]
    )
    def get(self, request):
        try:
            if request.user.is_staff:
                reports = Report.objects.all().order_by('-created_at')
            else:
                reports = Report.objects.filter(user=request.user).order_by('-created_at')
            
            paginator = self.pagination_class()
            paginated_reports = paginator.paginate_queryset(reports, request)
            serializer = ReportSerializer(paginated_reports, many=True, context={'request': request})
            
            return paginator.get_paginated_response(serializer.data)
        except Exception as e:
            LoggerService.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='ERROR',
                table_name='Report',
                description=f"Error retrieving reports: {str(e)}"
            )
            return Response({"detail": "Error retrieving reports"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @extend_schema(
        summary='Create a new report',
        description='Generates a new report based on the provided parameters. The report can be of different types and formats.',
        request=ReportCreateSerializer,
        responses={
            200: ReportSerializer,
            400: OpenApiResponse(description='Invalid input'),
            500: OpenApiResponse(description='Server error'),
        },
        examples=[
            OpenApiExample(
                name='Sales by Customer Report',
                summary='Generate a sales by customer report',
                description='Creates a report showing sales data organized by customer',
                value={
                    "name": "Monthly Sales by Customer",
                    "report_type": "sales_by_customer",
                    "format": "pdf",
                    "start_date": "2025-03-26",
                    "end_date": "2025-04-26",
                    "language": "en"
                },
                request_only=True,
            ),
            OpenApiExample(
                name='Best Sellers Report',
                summary='Generate a best sellers report',
                description='Creates a report showing the best selling products',
                value={
                    "name": "Best Sellers Q1",
                    "report_type": "best_sellers",
                    "format": "excel",
                    "start_date": "2025-01-01",
                    "end_date": "2025-03-31",
                    "language": "es"
                },
                request_only=True,
            ),
            OpenApiExample(
                name='Sales by Period Report',
                summary='Generate a sales by period report',
                description='Creates a report showing sales data organized by time periods',
                value={
                    "name": "Weekly Sales Report",
                    "report_type": "sales_by_period",
                    "format": "pdf",
                    "start_date": "2025-04-19",
                    "end_date": "2025-04-26",
                    "language": "en"
                },
                request_only=True,
            ),
            OpenApiExample(
                name='Product Performance Report',
                summary='Generate a product performance report',
                description='Creates a report analyzing product performance including profit margins',
                value={
                    "name": "Product Performance Analysis",
                    "report_type": "product_performance",
                    "format": "excel",
                    "start_date": "2025-03-01",
                    "end_date": "2025-04-26",
                    "language": "es"
                },
                request_only=True,
            ),
            OpenApiExample(
                name='Inventory Status Report',
                summary='Generate an inventory status report',
                description='Creates a report showing current inventory levels and status',
                value={
                    "name": "Current Inventory Status",
                    "report_type": "inventory_status",
                    "format": "pdf",
                    "language": "en"
                },
                request_only=True,
            ),
            OpenApiExample(
                name='My Orders Report',
                summary='Generate a personal orders report',
                description='Creates a report showing all your orders with payment and delivery status',
                value={
                    "name": "My Purchase History",
                    "report_type": "my_orders",
                    "format": "pdf",
                    "start_date": "2025-01-01", # Optional for this report type, defaults will be used
                    "end_date": "2025-04-30",   # Optional for this report type, defaults will be used
                    "language": "en"
                },
                request_only=True,
            ),
            OpenApiExample(
                name='Order Receipt',
                summary='Generate an order receipt',
                description='Creates a detailed receipt for a specific order',
                value={
                    "name": "Order Receipt #1234",
                    "report_type": "order_receipt",
                    "format": "pdf",
                    "language": "en",
                    "order_id": 1234
                },
                request_only=True,
            ),
        ]
    )
    def post(self, request):
        with transaction.atomic():
            try:
                if not request.user.is_authenticated:
                    return Response({"detail": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)
                
                serializer = ReportCreateSerializer(data=request.data, context={'request': request})
                if not serializer.is_valid():
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
                report_obj = serializer.save() 
                
                data = None
                if report_obj.report_type == 'sales_by_customer':
                    if not request.user.is_staff:
                        return Response({"detail": "Staff permission required for this report type"}, status=status.HTTP_403_FORBIDDEN)
                    data = self.generate_sales_by_customer_data(report_obj)
                elif report_obj.report_type == 'best_sellers':
                    if not request.user.is_staff:
                        return Response({"detail": "Staff permission required for this report type"}, status=status.HTTP_403_FORBIDDEN)
                    data = self.generate_best_sellers_data(report_obj)
                elif report_obj.report_type == 'sales_by_period':
                    if not request.user.is_staff:
                        return Response({"detail": "Staff permission required for this report type"}, status=status.HTTP_403_FORBIDDEN)
                    data = self.generate_sales_by_period_data(report_obj)
                elif report_obj.report_type == 'product_performance':
                    if not request.user.is_staff:
                        return Response({"detail": "Staff permission required for this report type"}, status=status.HTTP_403_FORBIDDEN)
                    data = self.generate_product_performance_data(report_obj)
                elif report_obj.report_type == 'inventory_status':
                    if not request.user.is_staff:
                        return Response({"detail": "Staff permission required for this report type"}, status=status.HTTP_403_FORBIDDEN)
                    data = self.generate_inventory_status_data(report_obj)
                elif report_obj.report_type == 'my_orders':
                    data = self.generate_my_orders_data(report_obj)
                elif report_obj.report_type == 'order_receipt':
                    # order_id was validated by serializer, retrieve it from original request data for this logic
                    order_id = request.data.get('order_id') 
                    if not order_id: # Should be caught by serializer, but double check
                        return Response({"detail": "Order ID is required for order receipt"}, status=status.HTTP_400_BAD_REQUEST)
                    data = self.generate_order_receipt_data(report_obj, order_id)
                else:
                    return Response({"detail": "Unknown report type"}, status=status.HTTP_400_BAD_REQUEST)
                
                report_obj.report_data = data
                
                file_content_bytes = None
                file_extension = report_obj.format

                if report_obj.format == 'pdf':
                    file_content_bytes = self.generate_pdf_content(report_obj)
                elif report_obj.format == 'excel':
                    file_content_bytes = self.generate_excel_content(report_obj)
                    file_extension = 'xlsx' 
                elif report_obj.format == 'json':
                    file_content_bytes = json.dumps(report_obj.report_data, indent=4, default=str).encode('utf-8')
                elif report_obj.format == 'html':
                    file_content_bytes = self.generate_html_content(report_obj)
                
                if file_content_bytes:
                    filename = f"{report_obj.report_type}_{report_obj.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
                    report_obj.file_path.save(filename, ContentFile(file_content_bytes), save=False) # save=False, will be saved with report_obj.save()
                
                report_obj.save() # Saves report_data and file_path
                
                LoggerService.objects.create(
                    user=request.user,
                    action='CREATE',
                    table_name='Report',
                    description=f"Report '{report_obj.name}' of type '{report_obj.report_type}' created."
                )
                return Response(ReportSerializer(report_obj, context={'request': request}).data, status=status.HTTP_200_OK)
                
            except Exception as e:
                logger.error(f"Error creating report: {str(e)}", exc_info=True)
                LoggerService.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    action='ERROR',
                    table_name='Report',
                    description=f"Error creating report: {str(e)}"
                )
                return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_translated_text(self, report, text_en):
        translations = {
            'en': {
                'Order ID': 'Order ID', 'Customer': 'Customer', 'Total Amount': 'Total Amount',
                'Currency': 'Currency', 'Date': 'Date', 'Product ID': 'Product ID',
                'Product Name': 'Product Name', 'Quantity Sold': 'Quantity Sold', 'Revenue': 'Revenue',
                'Period': 'Period', 'Orders': 'Orders', 'Sales': 'Sales',
                'Profit': 'Profit', 'Profit Margin': 'Profit Margin', 'Current Stock': 'Current Stock',
                'Reorder Level': 'Reorder Level', 'Status': 'Status', 'Count': 'Count',
                'Low Stock': 'Low Stock', 'In Stock': 'In Stock', 'Out of Stock': 'Out of Stock',
                'Report Date': 'Report Date', 'Date Range': 'Date Range', 'Generated by': 'Generated by',
                'Page': 'Page', 'Summary': 'Summary', 'Sales by Customer': 'Sales by Customer',
                'Best Sellers': 'Best Sellers', 'Sales by Period': 'Sales by Period',
                'Product Performance': 'Product Performance', 'Inventory Status': 'Inventory Status',
                'My Orders': 'My Orders', 'All Orders': 'All Orders', 'Payment Status': 'Payment Status',
                'Delivery Status': 'Delivery Status', 'Pending': 'Pending', 'Processing': 'Processing',
                'Completed': 'Completed', 'Failed': 'Failed', 'Refunded': 'Refunded',
                'Shipped': 'Shipped', 'Out_for_delivery': 'Out for Delivery', 'Delivered': 'Delivered',
                'Returned': 'Returned', 'Total Spent': 'Total Spent', 'Cost': 'Cost',
                'Product Code': 'Product Code', 'USD': 'USD', 'BS': 'BS',
                'Order Receipt': 'Order Receipt', 'Order': 'Order',
                'Order Information': 'Order Information', 'Order Items': 'Order Items',
                'Order Summary': 'Order Summary', 'Order Date:': 'Order Date:', 'Customer:': 'Customer:',
                'Email:': 'Email:', 'Shipping Address:': 'Shipping Address:',
                'Payment Status:': 'Payment Status:', 'Delivery Status:': 'Delivery Status:',
                'Product': 'Product', 'Unit Price': 'Unit Price', 'Quantity': 'Quantity',
                'Total': 'Total', 'Subtotal:': 'Subtotal:', 'Shipping:': 'Shipping:',
                'Taxes:': 'Taxes:', 'Discount:': 'Discount:',
                'Average Order Value': 'Average Order Value', 'Total Customers': 'Total Customers',
                'Total Revenue': 'Total Revenue', 'Category': 'Category', 'Average Price': 'Average Price',
                'Total Products Shown': 'Total Products Shown', 'Average Period Revenue': 'Average Period Revenue',
                'Total Products': 'Total Products', 'Total Inventory Value': 'Total Inventory Value',
                'Low Stock Items': 'Low Stock Items', 'Out of Stock Items': 'Out of Stock Items',
                'Total Stock': 'Total Stock', 'Items': 'Items', 'Total Items': 'Total Items',
                'Uncategorized': 'Uncategorized', 'Name': 'Name', 'Stock': 'Stock'
            },
            'es': {
                'Order ID': 'ID de Orden', 'Customer': 'Cliente', 'Total Amount': 'Monto Total',
                'Currency': 'Moneda', 'Date': 'Fecha', 'Product ID': 'ID de Producto',
                'Product Name': 'Nombre del Producto', 'Quantity Sold': 'Cantidad Vendida', 'Revenue': 'Ingresos',
                'Period': 'Período', 'Orders': 'Órdenes', 'Sales': 'Ventas',
                'Profit': 'Ganancia', 'Profit Margin': 'Margen de Ganancia', 'Current Stock': 'Stock Actual',
                'Reorder Level': 'Nivel de Reorden', 'Status': 'Estado', 'Count': 'Cantidad',
                'Low Stock': 'Stock Bajo', 'In Stock': 'En Stock', 'Out of Stock': 'Agotado',
                'Report Date': 'Fecha del Informe', 'Date Range': 'Rango de Fechas', 'Generated by': 'Generado por',
                'Page': 'Página', 'Summary': 'Resumen', 'Sales by Customer': 'Ventas por Cliente',
                'Best Sellers': 'Más Vendidos', 'Sales by Period': 'Ventas por Período',
                'Product Performance': 'Rendimiento de Productos', 'Inventory Status': 'Estado de Inventario',
                'My Orders': 'Mis Órdenes', 'All Orders': 'Todas las Órdenes', 'Payment Status': 'Estado de Pago',
                'Delivery Status': 'Estado de Entrega', 'Pending': 'Pendiente', 'Processing': 'Procesando',
                'Completed': 'Completado', 'Failed': 'Fallido', 'Refunded': 'Reembolsado',
                'Shipped': 'Enviado', 'Out_for_delivery': 'En reparto', 'Delivered': 'Entregado',
                'Returned': 'Devuelto', 'Total Spent': 'Total Gastado', 'Cost': 'Costo',
                'Product Code': 'Código de Producto', 'USD': 'USD',
                'Order Receipt': 'Recibo de Orden', 'Order': 'Orden',
                'Order Information': 'Información de la Orden', 'Order Items': 'Artículos de la Orden',
                'Order Summary': 'Resumen de la Orden', 'Order Date:': 'Fecha de la Orden:', 'Customer:': 'Cliente:',
                'Email:': 'Correo Electrónico:', 'Shipping Address:': 'Dirección de Envío:',
                'Payment Status:': 'Estado del Pago:', 'Delivery Status:': 'Estado de la Entrega:',
                'Product': 'Producto', 'Unit Price': 'Precio Unitario', 'Quantity': 'Cantidad',
                'Total': 'Total', 'Subtotal:': 'Subtotal:', 'Shipping:': 'Envío:',
                'Taxes:': 'Impuestos:', 'Discount:': 'Descuento:',
                'Average Order Value': 'Valor Promedio de Orden', 'Total Customers': 'Total de Clientes',
                'Total Revenue': 'Ingresos Totales', 'Category': 'Categoría', 'Average Price': 'Precio Promedio',
                'Total Products Shown': 'Total de Productos Mostrados', 'Average Period Revenue': 'Ingreso Promedio del Período',
                'Total Products': 'Total de Productos', 'Total Inventory Value': 'Valor Total de Inventario',
                'Low Stock Items': 'Artículos con Stock Bajo', 'Out of Stock Items': 'Artículos Agotados',
                'Total Stock': 'Stock Total', 'Items': 'Artículos', 'Total Items': 'Total de Artículos',
                'Uncategorized': 'Sin categorizar', 'Name': 'Nombre', 'Stock': 'Existencias'
            }
        }
        
        lang = report.language if report and report.language in translations else 'en'
        return translations[lang].get(text_en, text_en)
        
    def generate_sales_by_customer_data(self, report):
        start_date = report.start_date or (datetime.now() - timedelta(days=30)).date()
        end_date = report.end_date or datetime.now().date()
        
        orders = Order.objects.filter(
            created_at__date__range=[start_date, end_date],
            payment__payment_status='completed'
        ).select_related('user', 'payment')
        
        headers = [
            self.get_translated_text(report, 'Order ID'),
            self.get_translated_text(report, 'Customer'),
            self.get_translated_text(report, 'Total Amount'),
            self.get_translated_text(report, 'Currency'),
            self.get_translated_text(report, 'Date'),
        ]
        
        rows = []
        customer_totals = {} 
        
        for order in orders:
            customer_name = f"{order.user.first_name} {order.user.last_name}" if order.user else "Guest"
            if customer_name not in customer_totals:
                customer_totals[customer_name] = {"USD": 0, "BS": 0}
            
            order_currency = order.currency
            if order_currency in customer_totals[customer_name]:
                customer_totals[customer_name][order_currency] += float(order.total_amount)
            else:
                logger.warning(f"Unexpected currency '{order_currency}' in order {order.id} for sales_by_customer report.")

            rows.append({
                headers[0]: order.id,
                headers[1]: customer_name,
                headers[2]: float(order.total_amount),
                headers[3]: order_currency, 
                headers[4]: order.created_at.strftime('%Y-%m-%d')
            })
        
        summary = []
        for customer, totals in sorted(customer_totals.items(), key=lambda x: sum(x[1].values()), reverse=True):
            for currency_code, amount in totals.items(): 
                if amount > 0:
                    summary.append({
                        self.get_translated_text(report, 'Customer'): customer,
                        self.get_translated_text(report, 'Total Amount'): amount,
                        self.get_translated_text(report, 'Currency'): currency_code
                    })
            
        return {
            'title': self.get_translated_text(report, 'Sales by Customer'),
            'date_range': f"{start_date} - {end_date}",
            'headers': headers,
            'rows': rows,
            'summary': summary
        }
    
    def generate_best_sellers_data(self, report):
        start_date = report.start_date or (datetime.now() - timedelta(days=30)).date()
        end_date = report.end_date or datetime.now().date()
        
        completed_orders = Order.objects.filter(
            created_at__date__range=[start_date, end_date],
            payment__payment_status='completed'
        )
        
        order_ids = completed_orders.values_list('id', flat=True)
        
        items = OrderItem.objects.filter(order_id__in=order_ids).select_related('order', 'product')
        
        headers = [
            self.get_translated_text(report, 'Product ID'),
            self.get_translated_text(report, 'Product Name'),
            self.get_translated_text(report, 'Quantity Sold'),
            self.get_translated_text(report, 'Revenue'),
            self.get_translated_text(report, 'Currency')
        ]
        
        product_stats = {} 
        
        for item in items:
            if not item.product:
                continue
                
            product_id = item.product.id
            product_name = item.product.name
            item_order_currency = item.order.currency 
            
            if product_id not in product_stats:
                product_stats[product_id] = {
                    'name': product_name,
                    'quantity': 0,
                    'revenue': {"USD": 0, "BS": 0} 
                }
            
            product_stats[product_id]['quantity'] += item.quantity
            
            unit_price = 0
            if hasattr(item, 'unit_price') and item.unit_price is not None:
                unit_price = item.unit_price
            elif hasattr(item.product, 'price') and item.product.price is not None:
                unit_price = item.product.price
            
            total_price = unit_price * item.quantity
            
            if item_order_currency in product_stats[product_id]['revenue']:
                product_stats[product_id]['revenue'][item_order_currency] += float(total_price)
            else:
                logger.warning(f"Unexpected currency '{item_order_currency}' in order {item.order.id} for best_sellers report, product {product_id}.")

        sorted_products = sorted(
            product_stats.items(), 
            key=lambda x: x[1]['quantity'], 
            reverse=True
        )
        
        rows = []
        for product_id, stats in sorted_products:
            if stats['quantity'] > 0:
                for currency_code in ['USD', 'BS']: 
                    if stats['revenue'][currency_code] > 0:
                        rows.append({
                            headers[0]: product_id,
                            headers[1]: stats['name'],
                            headers[2]: stats['quantity'],
                            headers[3]: stats['revenue'][currency_code],
                            headers[4]: currency_code
                        })
        
        if not rows and sorted_products: 
            for product_id, stats in sorted_products[:10]: 
                if stats['quantity'] > 0:
                    rows.append({
                        headers[0]: product_id,
                        headers[1]: stats['name'],
                        headers[2]: stats['quantity'],
                        headers[3]: 0, 
                        headers[4]: 'USD' 
                    })

        return {
            'title': self.get_translated_text(report, 'Best Sellers'),
            'date_range': f"{start_date} - {end_date}",
            'headers': headers,
            'rows': rows
        }

    def generate_sales_by_period_data(self, report):
        start_date = report.start_date or (datetime.now() - timedelta(days=30)).date()
        end_date = report.end_date or datetime.now().date()
        
        orders = Order.objects.filter(
            created_at__date__range=[start_date, end_date],
            payment__payment_status='completed'
        ).select_related('payment')
        
        headers = [
            self.get_translated_text(report, 'Period'),
            self.get_translated_text(report, 'Orders'),
            self.get_translated_text(report, 'Sales'),
            self.get_translated_text(report, 'Currency')
        ]
        
        daily_sales = {} 
        
        for order in orders:
            date_str = order.created_at.date().strftime('%Y-%m-%d')
            if date_str not in daily_sales:
                daily_sales[date_str] = {
                    'USD': {'count': 0, 'total': 0},
                    'BS': {'count': 0, 'total': 0}
                }
            
            order_currency = order.currency
            if order_currency in daily_sales[date_str]:
                daily_sales[date_str][order_currency]['count'] += 1
                daily_sales[date_str][order_currency]['total'] += float(order.total_amount)
            else:
                logger.warning(f"Unexpected currency '{order_currency}' in order {order.id} for sales_by_period report.")
        
        rows = []
        for date_str, currencies_data in sorted(daily_sales.items()):
            for currency_code, data_val in currencies_data.items(): 
                if data_val['count'] > 0:
                    rows.append({
                        headers[0]: date_str,
                        headers[1]: data_val['count'],
                        headers[2]: data_val['total'],
                        headers[3]: currency_code
                    })
        
        total_by_currency = {'USD': {'orders': 0, 'sales': 0}, 'BS': {'orders': 0, 'sales': 0}}
        for date_data_val in daily_sales.values():
            for currency_code, data_val in date_data_val.items(): 
                if currency_code in total_by_currency: 
                    total_by_currency[currency_code]['orders'] += data_val['count']
                    total_by_currency[currency_code]['sales'] += data_val['total']
        
        summary = []
        for currency_code, totals in total_by_currency.items():
            if totals['orders'] > 0:
                summary.append({
                    self.get_translated_text(report, 'Currency'): currency_code,
                    self.get_translated_text(report, 'Orders'): totals['orders'],
                    self.get_translated_text(report, 'Sales'): totals['sales']
                })
        
        return {
            'title': self.get_translated_text(report, 'Sales by Period'),
            'date_range': f"{start_date} - {end_date}",
            'headers': headers,
            'rows': rows,
            'summary': summary
        }
        
    def generate_product_performance_data(self, report):
        start_date = report.start_date or (datetime.now() - timedelta(days=30)).date()
        end_date = report.end_date or datetime.now().date()
        
        completed_orders = Order.objects.filter(
            created_at__date__range=[start_date, end_date],
            payment__payment_status='completed'
        )
        
        order_ids = completed_orders.values_list('id', flat=True)
        items = OrderItem.objects.filter(order_id__in=order_ids).select_related('product', 'order')
        
        headers = [
            self.get_translated_text(report, 'Product ID'),
            self.get_translated_text(report, 'Product Name'),
            self.get_translated_text(report, 'Quantity Sold'),
            self.get_translated_text(report, 'Revenue'),
            self.get_translated_text(report, 'Currency')
        ]
        
        product_performance = {} 
        
        for item in items:
            if not item.product:
                continue
                
            product_id = item.product.id
            product_name = item.product.name
            item_order_currency = item.order.currency
            
            if product_id not in product_performance:
                product_performance[product_id] = {
                    'name': product_name,
                    'quantity': 0,
                    'revenue': {"USD": 0, "BS": 0}, 
                }
            
            unit_price = 0
            if hasattr(item, 'unit_price') and item.unit_price is not None:
                unit_price = item.unit_price
            elif hasattr(item.product, 'price') and item.product.price is not None:
                unit_price = item.product.price

            total_price = unit_price * item.quantity
                
            product_performance[product_id]['quantity'] += item.quantity
            if item_order_currency in product_performance[product_id]['revenue']:
                product_performance[product_id]['revenue'][item_order_currency] += float(total_price)
            else:
                logger.warning(f"Unexpected currency '{item_order_currency}' in order {item.order.id} for product_performance report, product {product_id}.")

        rows = []
        for product_id, data_val in sorted(product_performance.items(), key=lambda x: sum(x[1]['revenue'].values()), reverse=True):
            for currency_code in ['USD', 'BS']:
                if data_val['revenue'][currency_code] > 0:
                    rows.append({
                        headers[0]: product_id,
                        headers[1]: data_val['name'],
                        headers[2]: data_val['quantity'],
                        headers[3]: data_val['revenue'][currency_code],
                        headers[4]: currency_code
                    })
        
        return {
            'title': self.get_translated_text(report, 'Product Performance'),
            'date_range': f"{start_date} - {end_date}",
            'headers': headers,
            'rows': rows
        }

    def generate_inventory_status_data(self, report):
        inventory_items = Inventory.objects.all().select_related('product')
        
        headers = [
            self.get_translated_text(report, 'Product ID'),
            self.get_translated_text(report, 'Product Name'),
            self.get_translated_text(report, 'Current Stock'),
            self.get_translated_text(report, 'Status')
        ]
        
        rows = []
        status_counts = {
            'low_stock': 0,
            'in_stock': 0,
            'out_of_stock': 0
        }
        
        for inventory in inventory_items:
            if not inventory.product:
                continue
                
            current_stock = inventory.stock if hasattr(inventory, 'stock') else 0
            reorder_level = inventory.reorder_level if hasattr(inventory, 'reorder_level') else 10
            
            status_text = ''
            if current_stock <= 0:
                status_text = self.get_translated_text(report, 'Out of Stock')
                status_counts['out_of_stock'] += 1
            elif current_stock <= reorder_level:
                status_text = self.get_translated_text(report, 'Low Stock')
                status_counts['low_stock'] += 1
            else:
                status_text = self.get_translated_text(report, 'In Stock')
                status_counts['in_stock'] += 1
            
            rows.append({
                headers[0]: inventory.product.id,
                headers[1]: inventory.product.name,
                headers[2]: current_stock,
                headers[3]: status_text
            })
        
        rows.sort(key=lambda x: x[headers[2]])
        
        summary = [
            {
                self.get_translated_text(report, 'Status'): self.get_translated_text(report, 'Out of Stock'),
                self.get_translated_text(report, 'Count'): status_counts['out_of_stock']
            },
            {
                self.get_translated_text(report, 'Status'): self.get_translated_text(report, 'Low Stock'),
                self.get_translated_text(report, 'Count'): status_counts['low_stock']
            },
            {
                self.get_translated_text(report, 'Status'): self.get_translated_text(report, 'In Stock'),
                self.get_translated_text(report, 'Count'): status_counts['in_stock']
            }
        ]
        
        return {
            'title': self.get_translated_text(report, 'Inventory Status'),
            'headers': headers,
            'rows': rows,
            'summary': summary
        }

    def generate_my_orders_data(self, report):
        start_date = report.start_date or (datetime.now() - timedelta(days=90)).date()
        end_date = report.end_date or datetime.now().date()
        
        title_text = ''
        if report.user.is_staff:
            orders = Order.objects.filter(
                created_at__date__range=[start_date, end_date],
            ).select_related('payment', 'user', 'delivery').order_by('-created_at') 
            title_text = self.get_translated_text(report, 'All Orders')
        else:
            orders = Order.objects.filter(
                user=report.user,
                created_at__date__range=[start_date, end_date],
            ).select_related('payment', 'delivery').order_by('-created_at') 
            title_text = self.get_translated_text(report, 'My Orders')
        
        headers = [
            self.get_translated_text(report, 'Order ID'),
            self.get_translated_text(report, 'Date'),
        ]
        
        if report.user.is_staff:
            headers.insert(1, self.get_translated_text(report, 'Customer'))
            
        headers.extend([
            self.get_translated_text(report, 'Total Amount'),
            self.get_translated_text(report, 'Currency'),
            self.get_translated_text(report, 'Payment Status'),
            self.get_translated_text(report, 'Delivery Status')
        ])
        
        rows = []
        total_by_currency = {'USD': 0, 'BS': 0} 
        customer_totals = {} 
        
        for order in orders:
            payment_status_val = 'pending'
            delivery_status_val = 'pending'
            
            if hasattr(order, 'payment') and order.payment:
                payment_status_val = order.payment.payment_status
                
            if hasattr(order, 'delivery') and order.delivery: 
                delivery_status_val = order.delivery.delivery_status
            
            order_currency = order.currency 

            row_data = {
                headers[0]: order.id,
                headers[1 if not report.user.is_staff else 2]: order.created_at.strftime('%Y-%m-%d %H:%M'),
                headers[2 if not report.user.is_staff else 3]: float(order.total_amount),
                headers[3 if not report.user.is_staff else 4]: order_currency, 
                headers[4 if not report.user.is_staff else 5]: self.get_translated_text(report, payment_status_val.capitalize()),
                headers[5 if not report.user.is_staff else 6]: self.get_translated_text(report, delivery_status_val.capitalize())
            }
            
            if report.user.is_staff:
                customer_name = f"{order.user.first_name} {order.user.last_name}" if order.user else "Guest"
                row_data[headers[1]] = customer_name 
                
                if customer_name not in customer_totals:
                    customer_totals[customer_name] = {"USD": 0, "BS": 0}
                
                if hasattr(order, 'payment') and order.payment.payment_status == 'completed':
                    if order_currency in customer_totals[customer_name]:
                        customer_totals[customer_name][order_currency] += float(order.total_amount)
                    else:
                        logger.warning(f"Unexpected currency '{order_currency}' in order {order.id} for my_orders (staff) customer totals.")
            
            rows.append(row_data)
            
            if hasattr(order, 'payment') and order.payment.payment_status == 'completed':
                if order_currency in total_by_currency:
                    total_by_currency[order_currency] += float(order.total_amount)
                else:
                     logger.warning(f"Unexpected currency '{order_currency}' in order {order.id} for my_orders overall totals.")
        
        summary = []
        
        if report.user.is_staff and customer_totals:
            for customer, totals in sorted(customer_totals.items(), key=lambda x: sum(x[1].values()), reverse=True):
                for currency_code, amount in totals.items(): 
                    if amount > 0:
                        summary.append({
                            self.get_translated_text(report, 'Customer'): customer,
                            self.get_translated_text(report, 'Total Amount'): amount,
                            self.get_translated_text(report, 'Currency'): currency_code
                        })
        
        for currency_code, total in total_by_currency.items(): 
            if total > 0:
                summary.append({
                    self.get_translated_text(report, 'Currency'): currency_code,
                    self.get_translated_text(report, 'Total Spent'): total
                })
        
        return {
            'title': title_text,
            'date_range': f"{start_date} - {end_date}",
            'headers': headers,
            'rows': rows,
            'summary': summary
        }
        
    def generate_order_receipt_data(self, report, order_id): 
        try:
            order = None
            if report.user.is_staff:
                order = Order.objects.get(id=order_id)
            else:
                order = Order.objects.get(id=order_id, user=report.user)
                
            order_items = OrderItem.objects.filter(order=order).select_related('product')
            
            customer_name = f"{order.user.first_name} {order.user.last_name}" if order.user else "Guest"
            customer_email = order.user.email if order.user else ""
            
            shipping_address_str = ""
            delivery_obj = None
            
            if hasattr(order, 'delivery') and order.delivery:
                delivery_obj = order.delivery
                address_parts = []
                for field in ['address', 'city', 'state', 'country', 'postal_code']:
                    if hasattr(delivery_obj, field) and getattr(delivery_obj, field):
                        address_parts.append(str(getattr(delivery_obj, field)))
                shipping_address_str = ", ".join(address_parts)
            
            payment_status_val = "Pending"
            if hasattr(order, 'payment') and order.payment:
                payment_status_val = order.payment.payment_status.capitalize()
                
            delivery_status_val = "Pending"
            if delivery_obj and hasattr(delivery_obj, 'delivery_status'):
                delivery_status_val = delivery_obj.delivery_status.capitalize()
                
            order_info = {
                'order_id': order.id,
                'order_date': order.created_at.strftime('%Y-%m-%d %H:%M'),
                'customer': customer_name,
                'email': customer_email,
                'shipping_address': shipping_address_str,
                'payment_status': self.get_translated_text(report, payment_status_val),
                'delivery_status': self.get_translated_text(report, delivery_status_val),
                'subtotal': float(order.total_amount), 
                'shipping_fee': 0,
                'taxes': 0,
                'discount': 0,
                'total': float(order.total_amount),
                'currency': order.currency
            }
            
            if hasattr(order, 'subtotal') and order.subtotal is not None: 
                order_info['subtotal'] = float(order.subtotal)
            if hasattr(order, 'shipping_fee') and order.shipping_fee is not None:
                order_info['shipping_fee'] = float(order.shipping_fee)
            if hasattr(order, 'tax_amount') and order.tax_amount is not None:
                order_info['taxes'] = float(order.tax_amount)
            if hasattr(order, 'discount_amount') and order.discount_amount is not None:
                order_info['discount'] = float(order.discount_amount)
            
            if hasattr(order, 'subtotal') and order.subtotal is not None:
                 order_info['total'] = (order_info['subtotal'] + 
                                       order_info['shipping_fee'] + 
                                       order_info['taxes'] - 
                                       order_info['discount'])

            items_headers = [
                self.get_translated_text(report, 'Product'),
                self.get_translated_text(report, 'Unit Price'), 
                self.get_translated_text(report, 'Quantity'),
                self.get_translated_text(report, 'Total')
            ]
            
            items_rows = []
            for item in order_items:
                product_name = item.product.name if item.product else f"Product {item.product_id}"
                
                unit_price_val = 0
                if hasattr(item, 'unit_price') and item.unit_price is not None:
                    unit_price_val = float(item.unit_price)
                elif item.product and hasattr(item.product, 'price') and item.product.price is not None:
                    unit_price_val = float(item.product.price)

                quantity_val = item.quantity or 1 
                total_price_val = unit_price_val * quantity_val
                
                items_rows.append({
                    items_headers[0]: product_name,
                    items_headers[1]: unit_price_val, 
                    items_headers[2]: quantity_val,
                    items_headers[3]: float(total_price_val)
                })
                
            return {
                'title': self.get_translated_text(report, 'Order Receipt'),
                'subtitle': f"{self.get_translated_text(report, 'Order')} #{order.id}",
                'order': order_info,
                'items_headers': items_headers,
                'items': items_rows
            }
        except Order.DoesNotExist:
            raise Exception("Order not found or you don't have permission to access this order")
        
    def generate_pdf_content(self, report_obj):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles["h1"] 
        subtitle_style = styles["h2"]
        normal_style = styles["Normal"]
        bold_style = styles["h4"] 
        
        title = report_obj.report_data.get('title', report_obj.name)
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))
        
        if report_obj.report_type == 'order_receipt':
            subtitle = report_obj.report_data.get('subtitle', '')
            if subtitle:
                elements.append(Paragraph(subtitle, subtitle_style))
                elements.append(Spacer(1, 12))
            
            order = report_obj.report_data.get('order', {})
            if order:
                elements.append(Paragraph(self.get_translated_text(report_obj, 'Order Information'), bold_style))
                elements.append(Spacer(1, 8))
                
                customer_info_data = [
                    [Paragraph(self.get_translated_text(report_obj, 'Order Date:'), normal_style), Paragraph(order.get('order_date', ''), normal_style)],
                    [Paragraph(self.get_translated_text(report_obj, 'Customer:'), normal_style), Paragraph(order.get('customer', ''), normal_style)],
                    [Paragraph(self.get_translated_text(report_obj, 'Email:'), normal_style), Paragraph(order.get('email', ''), normal_style)],
                ]
                
                if order.get('shipping_address'):
                    customer_info_data.append([Paragraph(self.get_translated_text(report_obj, 'Shipping Address:'), normal_style), 
                                        Paragraph(order.get('shipping_address', ''), normal_style)])
                    
                customer_table = Table(customer_info_data, colWidths=[120, None]) 
                customer_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ]))
                elements.append(customer_table)
                elements.append(Spacer(1, 12))
                
                status_info_data = [
                    [Paragraph(self.get_translated_text(report_obj, 'Payment Status:'), normal_style), Paragraph(order.get('payment_status', ''), normal_style)],
                    [Paragraph(self.get_translated_text(report_obj, 'Delivery Status:'), normal_style), Paragraph(order.get('delivery_status', ''), normal_style)],
                ]
                
                status_table = Table(status_info_data, colWidths=[120, None])
                status_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ]))
                elements.append(status_table)
                elements.append(Spacer(1, 20))
                
                elements.append(Paragraph(self.get_translated_text(report_obj, 'Order Items'), bold_style))
                elements.append(Spacer(1, 8))
                
                items_headers = report_obj.report_data.get('items_headers', [])
                items = report_obj.report_data.get('items', [])
                
                if items_headers and items:
                    data = [[Paragraph(h, normal_style) for h in items_headers]] 
                    for item_row in items:
                        data.append([Paragraph(str(item_row.get(h, "")), normal_style) for h in items_headers])
                    
                    items_table = Table(data, repeatRows=1)
                    items_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('ALIGN', (1, 1), (1, -1), 'RIGHT'), 
                        ('ALIGN', (3, 1), (3, -1), 'RIGHT'), 
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ]))
                    elements.append(items_table)
                    elements.append(Spacer(1, 20))
                
                elements.append(Paragraph(self.get_translated_text(report_obj, 'Order Summary'), bold_style))
                elements.append(Spacer(1, 8))
                
                currency = order.get('currency', '')
                summary_info_data = [
                    [Paragraph(self.get_translated_text(report_obj, 'Subtotal:'), normal_style), Paragraph(f"{order.get('subtotal', 0):.2f} {currency}", normal_style)],
                ]
                
                if order.get('shipping_fee', 0) > 0:
                    summary_info_data.append([Paragraph(self.get_translated_text(report_obj, 'Shipping:'), normal_style), 
                                      Paragraph(f"{order.get('shipping_fee', 0):.2f} {currency}", normal_style)])
                
                if order.get('taxes', 0) > 0:
                    summary_info_data.append([Paragraph(self.get_translated_text(report_obj, 'Taxes:'), normal_style), 
                                      Paragraph(f"{order.get('taxes', 0):.2f} {currency}", normal_style)])
                
                if order.get('discount', 0) > 0:
                    summary_info_data.append([Paragraph(self.get_translated_text(report_obj, 'Discount:'), normal_style), 
                                      Paragraph(f"-{order.get('discount', 0):.2f} {currency}", normal_style)])
                
                summary_info_data.append([Paragraph(self.get_translated_text(report_obj, 'Total:'), bold_style), 
                                  Paragraph(f"{order.get('total', 0):.2f} {currency}", bold_style)]) 
                
                summary_table = Table(summary_info_data, colWidths=[120, None])
                summary_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, -1), (1, -1), 'Helvetica-Bold'), 
                    ('LINEABOVE', (0, -1), (1, -1), 1, colors.black),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                elements.append(summary_table)
        else:
            date_range = report_obj.report_data.get('date_range')
            if date_range:
                date_subtitle = f"{self.get_translated_text(report_obj, 'Date Range')}: {date_range}"
                elements.append(Paragraph(date_subtitle, subtitle_style))
                elements.append(Spacer(1, 12))
            
            headers = report_obj.report_data.get('headers', [])
            rows_data = report_obj.report_data.get('rows', [])
            
            if headers and rows_data:
                data = [[Paragraph(h, normal_style) for h in headers]]
                for row_dict in rows_data:
                    data.append([Paragraph(str(row_dict.get(h, "")), normal_style) for h in headers])
                
                table = Table(data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10), 
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 24))
            
            summary_data_list = report_obj.report_data.get('summary', [])
            if summary_data_list:
                elements.append(Paragraph(self.get_translated_text(report_obj, 'Summary'), subtitle_style))
                elements.append(Spacer(1, 12))
                
                summary_headers_set = set()
                for item in summary_data_list:
                    summary_headers_set.update(item.keys())
                summary_headers_list = list(summary_headers_set) 
                
                summary_table_data = [[Paragraph(h, normal_style) for h in summary_headers_list]]
                for item in summary_data_list:
                    summary_table_data.append([Paragraph(str(item.get(h, "")), normal_style) for h in summary_headers_list])
                
                summary_table = Table(summary_table_data, repeatRows=1)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                elements.append(summary_table)
        
        elements.append(Spacer(1, 48))
        footer_text_val = f"{self.get_translated_text(report_obj, 'Report Date')}: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if report_obj.user:
            footer_text_val += f", {self.get_translated_text(report_obj, 'Generated by')}: {report_obj.user.first_name} {report_obj.user.last_name}"
        elements.append(Paragraph(footer_text_val, normal_style))
        
        doc.build(elements, onFirstPage=self._header_footer, onLaterPages=self._header_footer)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
    
    def _header_footer(self, canvas, doc):
        canvas.saveState()
        
        header_text_val = "FICCT E-Commerce" 
        canvas.setFont('Helvetica-Bold', 16)
        canvas.drawString(doc.leftMargin, doc.height + doc.topMargin - 50, header_text_val) 
        
        canvas.setFont('Helvetica', 9)
        page_text_val = f"{self.get_translated_text(None, 'Page')} {doc.page}" 
        canvas.drawString(doc.leftMargin, doc.bottomMargin - 20, page_text_val) 
        
        canvas.restoreState()
        
    def generate_excel_content(self, report_obj):
        wb = Workbook()
        ws = wb.active
        ws.title = report_obj.report_type[:31] 
        
        header_font = Font(bold=True)
        title_font = Font(size=16, bold=True)
        subtitle_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        current_row = 1
        ws.cell(row=current_row, column=1, value=report_obj.report_data.get('title', report_obj.name)).font = title_font
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5) 
        current_row += 1
        
        if report_obj.report_type == 'order_receipt':
            subtitle = report_obj.report_data.get('subtitle', '')
            if subtitle:
                ws.cell(row=current_row, column=1, value=subtitle).font = subtitle_font
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            current_row +=1 

            order = report_obj.report_data.get('order', {})
            if order:
                ws.cell(row=current_row, column=1, value=self.get_translated_text(report_obj, 'Order Information')).font = subtitle_font
                current_row += 1
                
                info_pairs = [
                    (self.get_translated_text(report_obj, 'Order Date:'), order.get('order_date', '')),
                    (self.get_translated_text(report_obj, 'Customer:'), order.get('customer', '')),
                    (self.get_translated_text(report_obj, 'Email:'), order.get('email', '')),
                ]
                if order.get('shipping_address'):
                    info_pairs.append((self.get_translated_text(report_obj, 'Shipping Address:'), order.get('shipping_address', '')))
                info_pairs.extend([
                    (self.get_translated_text(report_obj, 'Payment Status:'), order.get('payment_status', '')),
                    (self.get_translated_text(report_obj, 'Delivery Status:'), order.get('delivery_status', '')),
                ])

                for label, value in info_pairs:
                    ws.cell(row=current_row, column=1, value=label).font = header_font
                    ws.cell(row=current_row, column=2, value=value)
                    current_row += 1
                current_row += 1 

                ws.cell(row=current_row, column=1, value=self.get_translated_text(report_obj, 'Order Items')).font = subtitle_font
                current_row += 1
                
                items_headers = report_obj.report_data.get('items_headers', [])
                items_data = report_obj.report_data.get('items', [])
                
                if items_headers and items_data:
                    for col_idx, header_text in enumerate(items_headers, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                        cell.font = header_font
                        cell.fill = header_fill
                    current_row += 1
                    for item_row_data in items_data:
                        for col_idx, header_text in enumerate(items_headers, 1):
                            ws.cell(row=current_row, column=col_idx, value=item_row_data.get(header_text, ""))
                        current_row += 1
                current_row += 1 

                ws.cell(row=current_row, column=1, value=self.get_translated_text(report_obj, 'Order Summary')).font = subtitle_font
                current_row += 1
                
                currency = order.get('currency', '')
                summary_pairs = [
                    (self.get_translated_text(report_obj, 'Subtotal:'), f"{order.get('subtotal', 0):.2f} {currency}"),
                ]
                if order.get('shipping_fee', 0) > 0:
                    summary_pairs.append((self.get_translated_text(report_obj, 'Shipping:'), f"{order.get('shipping_fee', 0):.2f} {currency}"))
                if order.get('taxes', 0) > 0:
                    summary_pairs.append((self.get_translated_text(report_obj, 'Taxes:'), f"{order.get('taxes', 0):.2f} {currency}"))
                if order.get('discount', 0) > 0:
                    summary_pairs.append((self.get_translated_text(report_obj, 'Discount:'), f"-{order.get('discount', 0):.2f} {currency}"))
                summary_pairs.append((self.get_translated_text(report_obj, 'Total:'), f"{order.get('total', 0):.2f} {currency}"))

                for label, value in summary_pairs:
                    ws.cell(row=current_row, column=1, value=label).font = header_font if label == self.get_translated_text(report_obj, 'Total:') else None
                    ws.cell(row=current_row, column=2, value=value).font = header_font if label == self.get_translated_text(report_obj, 'Total:') else None
                    current_row += 1
        else:
            date_range = report_obj.report_data.get('date_range')
            if date_range:
                ws.cell(row=current_row, column=1, value=f"{self.get_translated_text(report_obj, 'Date Range')}: {date_range}")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            current_row +=1 
            
            headers = report_obj.report_data.get('headers', [])
            rows_data = report_obj.report_data.get('rows', [])
            
            if headers and rows_data:
                for col_idx, header_text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                    cell.font = header_font
                    cell.fill = header_fill
                current_row += 1
                for row_dict_data in rows_data:
                    for col_idx, header_text in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col_idx, value=row_dict_data.get(header_text, ""))
                    current_row += 1
            current_row += 1 
            
            summary_list_data = report_obj.report_data.get('summary', [])
            if summary_list_data:
                ws.cell(row=current_row, column=1, value=self.get_translated_text(report_obj, 'Summary')).font = subtitle_font
                current_row += 1
                
                summary_headers_set = set()
                for item_dict_data in summary_list_data:
                    summary_headers_set.update(item_dict_data.keys())
                summary_headers_list_data = list(summary_headers_set) 

                if summary_headers_list_data:
                    for col_idx, header_text in enumerate(summary_headers_list_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                        cell.font = header_font
                        cell.fill = header_fill
                    current_row += 1
                    for item_dict_data in summary_list_data:
                        for col_idx, header_text in enumerate(summary_headers_list_data, 1):
                            ws.cell(row=current_row, column=col_idx, value=item_dict_data.get(header_text, ""))
                        current_row += 1
        
        current_row += 1 
        footer_text_val = f"{self.get_translated_text(report_obj, 'Report Date')}: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if report_obj.user:
            footer_text_val += f", {self.get_translated_text(report_obj, 'Generated by')}: {report_obj.user.first_name} {report_obj.user.last_name}"
        ws.cell(row=current_row, column=1, value=footer_text_val)
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        
        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes

    def generate_html_content(self, report_obj):
        html = f"""<!DOCTYPE html>
<html lang="{report_obj.language if report_obj else 'en'}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_obj.name if report_obj else 'Report'}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; color: #333; }}
        .container {{ max-width: 800px; margin: auto; padding: 20px; border: 1px solid #eee; box-shadow: 0 0 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 30px; border-bottom: 1px solid #eee; padding-bottom: 20px; }}
        .header h1 {{ margin-bottom: 5px; color: #333; }}
        .header p {{ color: #666; font-size: 0.9em; }}
        h2 {{ color: #333; border-bottom: 1px solid #eee; padding-bottom: 10px; margin-top: 30px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 0.9em; }}
        th {{ background-color: #f8f8f8; text-align: left; padding: 10px; border: 1px solid #ddd; }}
        td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
        tr:nth-child(even) {{ background-color: #fdfdfd; }}
        .summary-table th {{ text-align: right; font-weight: bold; }}
        .summary-table td {{ text-align: right; }}
        .total-row th, .total-row td {{ font-weight: bold; border-top: 2px solid #333; }}
        .footer {{ margin-top: 40px; font-size: 0.8em; color: #777; border-top: 1px solid #eee; padding-top: 10px; text-align: center; }}
        .info-table th {{ width: 30%; font-weight: bold; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{report_obj.report_data.get('title', report_obj.name if report_obj else 'Report')}</h1>"""
    
        if report_obj and report_obj.report_type == 'order_receipt' and 'subtitle' in report_obj.report_data:
             html += f"<p>{report_obj.report_data['subtitle']}</p>"
        elif report_obj and 'date_range' in report_obj.report_data and report_obj.report_data['date_range']:
            html += f"<p>{self.get_translated_text(report_obj, 'Date Range')}: {report_obj.report_data['date_range']}</p>"
    
        html += """
        </div>
        """
    
        if report_obj and report_obj.report_type == 'order_receipt':
            html += self._generate_order_receipt_html(report_obj)
        elif report_obj: 
            html += self._generate_standard_report_html(report_obj)
    
        html += f"""
        <div class="footer">
            <p>{self.get_translated_text(report_obj, 'Report Date')}: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>"""
        if report_obj and report_obj.user:
            html += f"""
            <p>{self.get_translated_text(report_obj, 'Generated by')}: {report_obj.user.first_name} {report_obj.user.last_name}</p>"""
        html += """
        </div>
    </div>
</body>
</html>"""
    
        return html.encode('utf-8')

    def _generate_standard_report_html(self, report_obj):
        html_content = ""
        report_data = report_obj.report_data or {}
        headers = report_data.get('headers', [])
        rows_data = report_data.get('rows', [])
    
        if headers and rows_data:
            html_content += """
        <h2>""" + self.get_translated_text(report_obj, report_data.get('title', 'Details')) + """</h2>
        <table>
            <thead>
                <tr>"""
            for header in headers:
                html_content += f"""
                    <th>{header}</th>"""
            html_content += """
                </tr>
            </thead>
            <tbody>"""
        
            for row_item in rows_data:
                html_content += """
                <tr>"""
                for header in headers:
                    value = row_item.get(header, "")
                    html_content += f"""
                    <td>{value}</td>"""
                html_content += """
                </tr>"""
        
            html_content += """
            </tbody>
        </table>"""
    
        summary_list = report_data.get('summary', [])
        if summary_list:
            html_content += f"""
        <h2>{self.get_translated_text(report_obj, 'Summary')}</h2>
        <table class="summary-table">"""
        
            if isinstance(summary_list, list) and all(isinstance(item, dict) for item in summary_list):
                summary_headers_set = set()
                for item in summary_list:
                    summary_headers_set.update(item.keys())
                summary_headers_list_data = list(summary_headers_set)

                if summary_headers_list_data:
                    html_content += "<thead><tr>"
                    for header in summary_headers_list_data:
                        html_content += f"<th>{header}</th>"
                    html_content += "</tr></thead><tbody>"
                    for item_dict in summary_list:
                        html_content += "<tr>"
                        for header in summary_headers_list_data:
                            html_content += f"<td>{item_dict.get(header, '')}</td>"
                        html_content += "</tr>"
                    html_content += "</tbody>"
            elif isinstance(summary_list, dict): 
                for key, value in summary_list.items():
                    html_content += f"""
                <tr>
                    <th>{key}</th>
                    <td>{value}</td>
                </tr>"""

            html_content += """
        </table>"""
    
        return html_content

    def _generate_order_receipt_html(self, report_obj):
        html_content = ""
        report_data = report_obj.report_data or {}
        order = report_data.get('order', {})
    
        if order:
            html_content += """
        <h2>""" + self.get_translated_text(report_obj, 'Order Information') + """</h2>
        <table class="info-table">
            <tbody>"""
        
            customer_info_pairs = [
                (self.get_translated_text(report_obj, 'Order Date:'), order.get('order_date', '')),
                (self.get_translated_text(report_obj, 'Customer:'), order.get('customer', '')),
                (self.get_translated_text(report_obj, 'Email:'), order.get('email', ''))
            ]
        
            if order.get('shipping_address'):
                customer_info_pairs.append((self.get_translated_text(report_obj, 'Shipping Address:'), order.get('shipping_address', '')))
            
            customer_info_pairs.extend([
                (self.get_translated_text(report_obj, 'Payment Status:'), order.get('payment_status', '')),
                (self.get_translated_text(report_obj, 'Delivery Status:'), order.get('delivery_status', ''))
            ])

            for label, value in customer_info_pairs:
                html_content += f"""
                <tr>
                    <th>{label}</th>
                    <td>{value}</td>
                </tr>"""
            html_content += """
            </tbody>
        </table>"""
        
            items_headers = report_data.get('items_headers', [])
            items_data = report_data.get('items', [])
        
            if items_headers and items_data:
                html_content += f"""
        <h2>{self.get_translated_text(report_obj, 'Order Items')}</h2>
        <table>
            <thead>
                <tr>"""
            
                for header in items_headers: 
                    html_content += f"""
                    <th>{header}</th>"""
            
                html_content += """
                </tr>
            </thead>
            <tbody>"""
            
                for item_row_data in items_data:
                    html_content += """
                    <tr>"""
                    for header in items_headers:
                        value = item_row_data.get(header, "")
                        try:
                            if header == self.get_translated_text(report_obj, 'Unit Price') or header == self.get_translated_text(report_obj, 'Total'):
                                html_content += f"""<td style="text-align: right;">{float(value):.2f} {order.get('currency', '')}</td>"""
                            elif header == self.get_translated_text(report_obj, 'Quantity'):
                                html_content += f"""<td style="text-align: center;">{value}</td>"""
                            else:
                                html_content += f"""<td>{value}</td>"""
                        except (ValueError, TypeError): 
                             html_content += f"""<td>{value}</td>"""
                    html_content += """
                    </tr>"""
            
                html_content += """
            </tbody>
        </table>"""
        
            html_content += f"""
        <h2>{self.get_translated_text(report_obj, 'Order Summary')}</h2>
        <table class="summary-table">
            <tbody>
                <tr>
                    <th>{self.get_translated_text(report_obj, 'Subtotal:')}</th>
                    <td>{order.get('subtotal', 0):.2f} {order.get('currency', '')}</td>
                </tr>"""
        
            if order.get('shipping_fee', 0) > 0:
                html_content += f"""
                <tr>
                    <th>{self.get_translated_text(report_obj, 'Shipping:')}</th>
                    <td>{order.get('shipping_fee', 0):.2f} {order.get('currency', '')}</td>
                </tr>"""
        
            if order.get('taxes', 0) > 0:
                html_content += f"""
                <tr>
                    <th>{self.get_translated_text(report_obj, 'Taxes:')}</th>
                    <td>{order.get('taxes', 0):.2f} {order.get('currency', '')}</td>
                </tr>"""
        
            if order.get('discount', 0) > 0:
                html_content += f"""
                <tr>
                    <th>{self.get_translated_text(report_obj, 'Discount:')}</th>
                    <td>-{order.get('discount', 0):.2f} {order.get('currency', '')}</td>
                </tr>"""
        
            html_content += f"""
                <tr class="total-row">
                    <th>{self.get_translated_text(report_obj, 'Total:')}</th>
                    <td>{order.get('total', 0):.2f} {order.get('currency', '')}</td>
                </tr>
            </tbody>
        </table>"""
    
        return html_content